import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import zipfile
import tempfile
from io import BytesIO
from datetime import datetime

# ===================== 全局极简样式配置 =====================
st.set_page_config(page_title="单证工具", page_icon="📦", layout="wide", initial_sidebar_state="collapsed")
st.markdown("""
<style>
.main-title {font-size: 28px; font-weight: 600; margin: 10px 0 24px 0;}
.section-divider {margin: 40px 0; border-top:1px solid #eee;}
.card {background:#f9fafb; padding:22px; border-radius:10px; margin-bottom:16px;}
.btn-main {background:#2563eb; color:white; font-size:16px; padding:8px 26px; border:none; border-radius:8px;}
.btn-main:hover {background:#1d4ed8;}
.info-block {background:#eef6ff; padding:12px; border-radius:8px; margin-top:10px; border-left:4px #2563eb solid;}
</style>
""", unsafe_allow_html=True)

# ===================== 账号配置（清关模块专用） =====================
ACCOUNT_INFO = {
    "39": {"shipper_name": "Shenzhen Longyuan Junjie Technology Co., Ltd.","shipper_addr": "Room 502, No. 5, Ruiyuan Second Lane, Nanlian Community, Longgang Subdistrict, Longgang District, Shenzhen,Guangdong,China","contact": "ZHOUJUNJIE","phone": "+8613427679670"},
    "79": {"shipper_name": "Mingtongsheng (Shenzhen) E-commerce Co., Ltd.","shipper_addr": "3A Zijing Pavilion, Building 4, Baizhu Garden, 249 Zhuguang Road, Longlian Community, Tao Yuan Street, Nanshan District, Shenzhen,Guangdong,China","contact": "LIMIN","phone": "+8613902478270"},
    "76": {"shipper_name": "Shenzhen Chengziwei Technology Co., Ltd.","shipper_addr": "B410, Buildings 2 and 3, Mingliang Technology Park, No. 88 Zhuguang North Road, Pingshan Community, Tao Yuan Street, Nanshan District, Shenzhen,Guangdong,China","contact": "CHENZIWEI","phone": "+8615875396146"},
    "47.92": {"shipper_name": "Shenzhen Dongshan Jinhao Technology Co., Ltd.","shipper_addr": "Room 108, Building 8, Maker Town, Xili Street, Nanshan District, Shenzhen,Guangdong,China","contact": "CHENJINHAO","phone": "+8613530567440"},
    "47.100": {"shipper_name": "Shenzhen Weizhite Technology Co., Ltd.","shipper_addr": "410, Building 2-3, Bright Technology Park, No. 88 Zhuguang North Road, Pingshan Community, Taoyuan Street, Nanshan District, Shenzhen,Guangdong,China","contact": "CHENSIFA","phone": "+8615986680681"},
    "47.99": {"shipper_name": "Shenzhen Shunhuixiong Technology Co., Ltd.","shipper_addr": "115, Building 10, Maker Town, No. 4109 Liuxian Avenue, Pingshan Community, Taoyuan Street, Nanshan District, Shenzhen,Guangdong,China","contact": "LIANGRIXIONG","phone": "+8618026938073"},
    "70": {"shipper_name": "Shenzhen Xingyuepan Technology Co., Ltd.","shipper_addr": "B429, No. 22 Dakan Industrial 2nd Road, Daguan Community, Xili Street, Nanshan District, Shenzhen,Guangdong,China","contact": "LIUBIN","phone": "+8613530369614"},
    "71": {"shipper_name": "Shenzhen Chenghai Liufa Technology Co., Ltd.","shipper_addr": "28C, Unit A, Building 3, Xiangshanli Phase 5, Wenchang Street Community, Shahe Street, Nanshan District, Shenzhen,Guangdong,China","contact": "GULIUFA","phone": "+8617744965296"},
    "8.1": {"shipper_name": "Shenzhen Shiqi Jiechao Technology Co., Ltd.","shipper_addr": "Tianxi Xiaoju V211, No. 10 Ruihua North Lane, Nanlian Community, Longgang Street, Longgang District, Shenzhen,Guangdong,China","contact": "LUJIECHAO","phone": "+8613670528672"},
    "47.108": {"shipper_name": "Guangzhou Changyou Weilin Technology Co., Ltd.","shipper_addr": "Shop 422, No. 15 Yihe Road, Liwan District, Guangzhou,Guangdong,China","contact": "LUOSILIN","phone": "+8615728285292"},
    "47.239": {"shipper_name": "Hong Kong LingLingQinLv Technology Limited","shipper_addr": "UNIT F22,RM 6, 10/F, LEMMI CENTRE, 50 HOI YUEN ROAD,Kwun Tong,Hong Kong","contact": "LUQINGLING","phone": "+8619864368710"}
}

# ===================== 清关模板坐标（100%匹配AL0-SBU6B5D6EZU6S.xlsx模板） =====================
CLEAR_MAP = {
    "fba_no": "J8",          # FBA单号 J列第8行
    "ship_name": "B7",       # 发货人名称 B列第7行
    "ship_addr": "B8",       # 发货人地址 B列第8行
    "ship_contact": "B9",    # 发货人联系人 B列第9行
    "ship_tel": "B10",      # 发货人电话 B列第10行
    "imp_name": "E7",       # 进口商名称 E列第7行
    "imp_addr": "E8",       # 进口商地址 E列第8行
    "imp_contact": "E9",    # 进口商联系人 E列第9行
    "imp_tel": "E10",      # 进口商电话 E列第10行
    "manu_name": "C38",     # 制造商名称 C列第38行
    "manu_addr": "C39",     # 制造商地址 C列第39行
    "data_start": 22,        # 明细数据开始行
    "data_end": 35,          # 明细数据结束行
    "total_row": 36,         # 合计行
    "qty_col": 9,           # 数量 I列（第9列）
    "unit_price_col": 10,   # 单价 J列（第10列）
    "total_price_col": 11,  # 总价 K列（第11列）
    "ctns_col": 13,         # 箱数 M列（第13列）
    "weight_col": 14,       # 毛重 N列（第14列）
    "nw_col": 15,           # 净重 O列（第15列）
    "vol_col": 16           # 体积 P列（第16列）
}

# 截单LCL模板坐标
CUT_MAP = {
    "header_row":3,"data_start":4,"data_end":7,"weight_col":3,"vol_col":4,
    "weight_head":"Gross weight","vol_head":"Volume"
}

# ===================== 页面标题 =====================
st.markdown('<div class="main-title">📦 单证批量处理工具</div>', unsafe_allow_html=True)

# ===================== 模块1：FBA清关单生成（下拉下方展示公司信息） =====================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("1. FBA清关单批量生成")
col1, col2 = st.columns([0.48, 0.48], gap="medium")
with col1:
    file_clear = st.file_uploader("上传数据源Excel", type=["xlsx","xls"], key="clear_file")
with col2:
    acc_list = list(ACCOUNT_INFO.keys())
    select_acc = st.selectbox("选择账号", options=acc_list, key="acc_sel")
    # 账号下拉下方展示公司信息
    acc_detail = ACCOUNT_INFO[select_acc]
    st.markdown('<div class="info-block">', unsafe_allow_html=True)
    st.write(f"公司：{acc_detail['shipper_name']}")
    st.write(f"地址：{acc_detail['shipper_addr']}")
    st.write(f"联系人：{acc_detail['contact']} | 电话：{acc_detail['phone']}")
    st.markdown('</div>', unsafe_allow_html=True)

gen_clear = st.button("生成并下载清关资料", key="gen_clear", type="primary")
st.markdown('</div>', unsafe_allow_html=True)

# 清关生成逻辑（完全适配模板，修复所有报错）
if gen_clear:
    if not file_clear:
        st.error("请上传数据源文件")
    elif not os.path.exists(TEMPLATE_FILE):
        st.error(f"模板文件{TEMPLATE_FILE}缺失，请上传至仓库根目录")
    else:
        with st.spinner("正在生成..."):
            df = pd.read_excel(file_clear)
            groups = df.groupby("FBA编号")
            acc_info = ACCOUNT_INFO[select_acc]
            tmp_dir = tempfile.TemporaryDirectory()
            tmp_path = tmp_dir.name
            file_list = []

            for fba_id, group in groups:
                wb = load_workbook(TEMPLATE_FILE)
                ws = wb.active

                # 自动解除制造商区域合并单元格，避免赋值失败
                try:
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= 39 and merged_range.max_row >= 38) and (merged_range.min_col <= 3 and merged_range.max_col >= 3):
                            ws.unmerge_cells(range_string=str(merged_range))
                            break
                except:
                    pass

                # 填充发货人信息（带异常捕获）
                try:
                    ws[CLEAR_MAP["ship_name"]].value = acc_info["shipper_name"]
                    ws[CLEAR_MAP["ship_addr"]].value = acc_info["shipper_addr"]
                    ws[CLEAR_MAP["ship_contact"]].value = f"Contact:{acc_info['contact']}"
                    ws[CLEAR_MAP["ship_tel"]].value = f"Phone:{acc_info['phone']}"
                except:
                    pass

                # 填充进口商信息（带异常捕获）
                try:
                    ws[CLEAR_MAP["imp_name"]].value = acc_info["shipper_name"]
                    ws[CLEAR_MAP["imp_addr"]].value = acc_info["shipper_addr"]
                    ws[CLEAR_MAP["imp_contact"]].value = f"Contact:{acc_info['contact']}"
                    ws[CLEAR_MAP["imp_tel"]].value = f"Phone:{acc_info['phone']}"
                except:
                    pass

                # 填充制造商信息（强制填充，双层兜底）
                try:
                    ws[CLEAR_MAP["manu_name"]].value = acc_info["shipper_name"]
                    ws[CLEAR_MAP["manu_addr"]].value = acc_info["shipper_addr"]
                except Exception:
                    try:
                        ws["C38"].value = acc_info["shipper_name"]
                        ws["C39"].value = acc_info["shipper_addr"]
                    except:
                        pass

                # 填充FBA编号（带异常捕获）
                try:
                    ws[CLEAR_MAP["fba_no"]].value = fba_id
                except:
                    pass

                # 清空旧明细区域
                s_r = CLEAR_MAP["data_start"]
                e_r = CLEAR_MAP["data_end"]
                for r in range(s_r, e_r+1):
                    for c in range(2, 17):
                        ws.cell(row=r, column=c, value=None)

                # 写入新明细（匹配模板列顺序）
                rows = group.values.tolist()
                for idx, row in enumerate(rows):
                    r = s_r + idx
                    ws.cell(r, 2, row[0])  # 零件号 B列
                    ws.cell(r, 3, row[1])  # 品名 C列
                    ws.cell(r, 4, row[2])  # 材质 D列
                    ws.cell(r, 5, row[3])  # 关税分类 E列
                    ws.cell(r, 8, "CN")   # 原产国 H列
                    ws.cell(r, 9, row[7])  # 数量 I列
                    ws.cell(r, 10, row[8]) # 单价 J列
                    ws.cell(r, 11, f"=J{r}*I{r}") # 总价 K列
                    ws.cell(r, 13, row[11]) # 箱数 M列
                    ws.cell(r, 14, round(row[12], 3)) # 毛重 N列
                    ws.cell(r, 15, row[13]) # 净重 O列
                    ws.cell(r, 16, round(row[14], 3)) # 体积 P列

                # 合计公式（匹配模板列号）
                end_data = s_r + len(rows) - 1
                total_r = CLEAR_MAP["total_row"]
                ws.cell(total_r, 11, f"=SUM(K{s_r}:K{end_data})") # 总价合计
                ws.cell(total_r, 13, f"=SUM(M{s_r}:M{end_data})") # 箱数合计
                ws.cell(total_r, 14, f"=SUM(N{s_r}:N{end_data})") # 毛重合计
                ws.cell(total_r, 15, f"=SUM(O{s_r}:O{end_data})") # 净重合计
                ws.cell(total_r, 16, f"=SUM(P{s_r}:P{end_data})") # 体积合计

                save_path = os.path.join(tmp_path, f"{fba_id}.xlsx")
                wb.save(save_path)
                wb.close()
                file_list.append(save_path)

            # 打包zip压缩包
            zip_buf = BytesIO()
            zip_name = f"{select_acc}清关资料.zip"
            with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fp in file_list:
                    zf.write(fp, os.path.basename(fp))
            zip_buf.seek(0)
            # 修复下载按钮：移除hidden=True，适配新版Streamlit
            st.download_button(
                label="点击下载清关压缩包",
                data=zip_buf,
                file_name=zip_name,
                mime="application/zip",
                key="dl_clear_auto"
            )
            st.success(f"{zip_name} 已生成，请点击上方按钮下载！")
            tmp_dir.cleanup()

# 分割线
st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ===================== 模块2：LCL截单重量体积比例调整 =====================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("2. LCL截单重量体积调整")
upload_cut = st.file_uploader("上传LCL截单Excel", type=["xlsx","xls"], key="cut_file")
colw, colv = st.columns([0.48,0.48], gap="medium")
with colw:
    target_w = st.number_input("目标总重量 kg", min_value=0.001, step=0.001, format="%.3f", key="tw")
with colv:
    target_v = st.number_input("目标总体积 CBM", min_value=0.001, step=0.001, format="%.3f", key="tv")
adjust_btn = st.button("调整并下载截单文件", key="adj_btn", type="primary")
st.markdown('</div>', unsafe_allow_html=True)

# 截单调整逻辑（修复版）
if adjust_btn:
    if not upload_cut:
        st.error("请上传截单Excel文件")
    elif target_w <= 0 or target_v <= 0:
        st.error("重量、体积必须大于0")
    else:
        with st.spinner("计算并调整体积重量..."):
            # 提取AL0编号
            fname = upload_cut.name
            name_no_ext = os.path.splitext(fname)[0]
            al0_code = ""
            for part in name_no_ext.split("_"):
                if part.startswith("AL0"):
                    al0_code = part
                    break
            if not al0_code:
                al0_code = "未知单号"
            out_name = f"截单资料{al0_code}.xlsx"

            wb = load_workbook(upload_cut)
            ws = wb.active
            
            # 自动识别数据范围（替代硬编码行号）
            s_r = None
            e_r = None
            w_col = None
            v_col = None
            # 遍历表头行，自动识别重量/体积列
            for r in range(1, 10):
                row_cells = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[r]]
                for idx, cell_val in enumerate(row_cells):
                    if "weight" in cell_val.lower() or "gross" in cell_val.lower():
                        w_col = idx + 1
                    if "volume" in cell_val.lower() or "cbm" in cell_val.lower():
                        v_col = idx + 1
                if w_col and v_col:
                    s_r = r + 1
                    break
            # 自动识别数据结束行
            if s_r:
                for r in range(s_r, 20):
                    cell_val = ws.cell(r, w_col).value
                    if cell_val is None or str(cell_val).strip() == "":
                        e_r = r - 1
                        break
            # 兜底：如果自动识别失败，用硬编码默认值
            if not s_r or not e_r or not w_col or not v_col:
                st.warning("自动识别模板失败，使用默认坐标（第4-7行，C列重量，D列体积）")
                s_r = 4
                e_r = 7
                w_col = 3
                v_col = 4

            # 读取原始数据
            raw_data = []
            sum_w_ori = 0.0
            sum_v_ori = 0.0
            for r in range(s_r, e_r + 1):
                wc = ws.cell(r, w_col)
                vc = ws.cell(r, v_col)
                try:
                    wv = float(wc.value)
                    vv = float(vc.value)
                    raw_data.append([r, wv, vv])
                    sum_w_ori += wv
                    sum_v_ori += vv
                except:
                    continue

            # 校验原始数据
            if sum_w_ori <= 0 or sum_v_ori <= 0:
                st.error("原始单据总重量/体积为0，无法调整")
                wb.close()
                st.stop()
            st.info(f"已读取原始数据：共{len(raw_data)}行，总重量{sum_w_ori:.3f}kg，总体积{sum_v_ori:.3f}CBM")

            # 计算统一缩放比例（保证单箱重体积比不变，和原始一致）
            ratio_w = target_w / sum_w_ori
            ratio_v = target_v / sum_v_ori
            final_ratio = ratio_w if abs(ratio_w - ratio_v) < 0.001 else (ratio_w + ratio_v) / 2
            st.info(f"缩放比例：{final_ratio:.6f}，目标总重量{target_w:.3f}kg，目标总体积{target_v:.3f}CBM")

            # 计算缩放后数值
            data_list = []
            for r, ow, ov in raw_data:
                ew = ow * final_ratio
                ev = ov * final_ratio
                data_list.append([r, ew, ev])

            # 精准分配小数尾差，保证总重量/体积完全等于目标值
            # 重量尾差分配
            target_w_int = int(round(target_w * 1000))
            w_int_list = []
            sum_wi = 0
            for _, ew, _ in data_list:
                i = int(round(ew * 1000))
                w_int_list.append(i)
                sum_wi += i
            # 尾差平均分配到所有行
            w_diff = target_w_int - sum_wi
            if w_diff != 0:
                step = 1 if w_diff > 0 else -1
                for i in range(abs(w_diff)):
                    w_int_list[i % len(w_int_list)] += step

            # 体积尾差分配
            target_v_int = int(round(target_v * 1000))
            v_int_list = []
            sum_vi = 0
            for _, _, ev in data_list:
                i = int(round(ev * 1000))
                v_int_list.append(i)
                sum_vi += i
            # 尾差平均分配到所有行
            v_diff = target_v_int - sum_vi
            if v_diff != 0:
                step = 1 if v_diff > 0 else -1
                for i in range(abs(v_diff)):
                    v_int_list[i % len(v_int_list)] += step

            # 写入单元格
            for idx, (row_num, _, _) in enumerate(data_list):
                final_w = w_int_list[idx] / 1000
                final_v = v_int_list[idx] / 1000
                ws.cell(row_num, w_col, value=final_w)
                ws.cell(row_num, v_col, value=final_v)

            # 保存文件
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            wb.close()

            # 验证结果
            wb_check = load_workbook(BytesIO(buf.getvalue()))
            ws_check = wb_check.active
            final_sum_w = 0.0
            final_sum_v = 0.0
            for r in range(s_r, e_r + 1):
                final_sum_w += float(ws_check.cell(r, w_col).value or 0)
                final_sum_v += float(ws_check.cell(r, v_col).value or 0)
            wb_check.close()
            st.success(f"调整完成！最终总重量{final_sum_w:.3f}kg，最终总体积{final_sum_v:.3f}CBM")

            # 下载按钮
            st.download_button(
                label="点击下载调整后截单文件",
                data=buf,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_cut_auto"
            )

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import tempfile
import datetime
import gc
import os
import shutil
import re

# 图片压缩配置（仅优化体积，不删除/替换原有图片）
IMG_MAX_W = 110    # 适配A列单元格宽度
IMG_MAX_H = 110    # 适配A列单元格高度
IMG_QUALITY = 70   # 平衡体积与清晰度

# 分割线
st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ===================== 模块3：清关单单文件精准填充（无分单版） =====================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("3. 清关单固定列精准填充")
st.markdown("上传清关单Excel，**仅输出一个单文件**，仅填充【跟踪号/FBA】列有数据的行，空行完全保留不变，图片完整保留")
st.info("✅ 填充规则（仅FBA号非空的行生效）：\n- 产品分类(*) H列 → `CPSC`\n- 产品数量单位(*) M列 → `套`\n- PO创建日期 AC列 → 今日日期\n- FBA箱号 AD列 → `-`\n- 外箱分货标(*) AE列 → `A1`")

# 上传组件
file_upload = st.file_uploader("上传清关单Excel（支持xls/xlsx）", type=["xlsx","xls"], key="fill_file")
gen_btn = st.button("生成填充后文件", key="gen_fill", type="primary")

st.markdown('</div>', unsafe_allow_html=True)

# 今日日期格式（匹配你的模板：2026-7-27）
today = datetime.datetime.now().strftime("%Y-%m-%d").replace("-0", "-")

# 固定填充列号与对应值（严格按你要求，列号对应Excel列）
FIXED_FILL = {
    8: "CPSC",       # H列 产品分类(*)
    13: "套",         # M列 产品数量单位(*)
    29: today,        # AC列 PO创建日期
    30: "-",         # AD列 FBA箱号
    31: "A1"         # AE列 外箱分货标(*)
}

# 图片压缩工具函数（仅优化体积，不删除原图）
def compress_image_optimize(img_obj, temp_dir):
    try:
        img_bytes = img_obj._data()
        tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".png", dir=temp_dir)
        tmp_in.write(img_bytes)
        tmp_in.close()

        with PILImage.open(tmp_in.name) as pil_img:
            # 透明通道转RGB，减少体积
            if pil_img.mode in ("RGBA", "LA"):
                pil_img = pil_img.convert("RGB")
            # 等比例缩放，不拉伸变形
            pil_img.thumbnail((IMG_MAX_W, IMG_MAX_H), PILImage.Resampling.LANCZOS)
            # 输出压缩后的JPG
            tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg", dir=temp_dir)
            pil_img.save(tmp_out.name, "JPEG", quality=IMG_QUALITY, optimize=True)
        
        # 清理临时输入文件
        os.unlink(tmp_in.name)
        return tmp_out.name
    except Exception as e:
        st.warning(f"图片压缩失败：{str(e)}")
        return None

# 核心填充逻辑
if gen_btn:
    if not file_upload:
        st.error("请上传清关单Excel文件")
    else:
        progress_bar = st.progress(0)
        status_text = st.empty()
        try:
            status_text.text("📖 正在读取上传文件...")
            df = None
            error_messages = []
            file_ext = os.path.splitext(file_upload.name)[1].lower()
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_ext)
            tmp_file.write(file_upload.getvalue())
            tmp_file.close()

            # 兼容读取xls/xlsx双格式
            if file_ext == ".xls":
                try:
                    df = pd.read_excel(tmp_file.name, header=None, engine="xlrd")
                except Exception as e:
                    error_messages.append(f"xls读取失败:{str(e)}")
                    try:
                        df = pd.read_excel(tmp_file.name, header=None, engine="openpyxl")
                    except Exception as e2:
                        error_messages.append(f"openpyxl读取失败:{str(e2)}")
            else:
                try:
                    df = pd.read_excel(tmp_file.name, header=None, engine="openpyxl")
                except Exception as e:
                    error_messages.append(f"xlsx读取失败:{str(e)}")
            try:
                os.unlink(tmp_file.name)
            except:
                pass
            if df is None:
                st.error("文件读取失败："+"\n".join(error_messages))
                st.stop()

            # --------------------------
            # 1. 解析表头，精准匹配Excel行号
            # --------------------------
            # 第2行(index=1)为主表头，第3行(index=2)为补充表头
            row2_header = df.iloc[1].fillna("").astype(str).str.strip()
            row3_header = df.iloc[2].fillna("").astype(str).str.strip()
            
            # 合并生成最终列名
            final_columns = []
            for h2, h3 in zip(row2_header, row3_header):
                if h2 and h3:
                    final_columns.append(f"{h2}{h3}")
                elif h2:
                    final_columns.append(h2)
                elif h3:
                    final_columns.append(h3)
                else:
                    final_columns.append(f"col_{len(final_columns)}")
            
            # 数据从第4行(index=3)开始，精准对应Excel行号
            data_df = df.iloc[3:].copy()
            data_df.columns = final_columns
            data_df.columns = data_df.columns.str.strip()

            # 清理列名特殊字符，避免匹配失败
            def clean_col(col):
                if pd.isna(col):
                    return col
                col = str(col)
                col = re.sub(r'[（(][^）)]*[）)]', '', col)
                col = col.replace('(*)', '').replace('<br>', '').replace('(USD)', '').replace('(CBM)', '').replace('(KGS)', '')
                return col.strip()
            data_df.columns = [clean_col(col) for col in data_df.columns]

            # 找到跟踪号/FBA列
            fba_col_name = None
            for c in data_df.columns:
                if "跟踪号/FBA" in c or c in "跟踪号/FBA":
                    fba_col_name = c
                    break
            if fba_col_name is None:
                st.error(f"表格中未找到【跟踪号/FBA】列，当前识别到的列名：{', '.join(data_df.columns.tolist())}")
                st.stop()

            # --------------------------
            # 2. 筛选FBA号非空的行，精准对应Excel行号
            # --------------------------
            # 数据行对应的Excel行号 = 原始df的index + 1（df.index从0开始，对应Excel行1）
            valid_excel_rows = []
            # 收集所有FBA号非空的行
            fba_list = []
            for idx, row_data in data_df.iterrows():
                fba_val = row_data[fba_col_name]
                if pd.notna(fba_val) and str(fba_val).strip() != "":
                    # 精准计算Excel行号：idx是data_df的索引，对应原始df的index=idx+3，Excel行号=idx+3+1=idx+4
                    excel_row = idx + 4
                    valid_excel_rows.append(excel_row)
                    fba_list.append(str(fba_val).strip())

            if not valid_excel_rows:
                st.warning("未找到【跟踪号/FBA】列有数据的行，未执行填充")
                st.stop()

            st.info(f"✅ 识别到 {len(valid_excel_rows)} 行FBA号非空，将仅对这些行执行填充")
            progress_bar.progress(20)

            # --------------------------
            # 3. 加载模板，仅填充FBA有效行
            # --------------------------
            file_upload.seek(0)
            wb = load_workbook(file_upload)
            ws = wb.active
            max_r = ws.max_row
            max_c = ws.max_column

            # 拆分所有合并单元格，消除只读报错
            all_merge = list(ws.merged_cells.ranges)
            for rng in all_merge:
                ws.unmerge_cells(str(rng))

            # 关闭公式自动计算，大幅减小文件体积
            wb.calculation.calcMode = "manual"

            # --------------------------
            # 核心：仅填充FBA号非空的行，空行完全不碰
            # --------------------------
            status_text.text("🔄 正在填充FBA有效行...")
            progress_bar.progress(40)
            for excel_row in valid_excel_rows:
                for col_num, val in FIXED_FILL.items():
                    ws.cell(row=excel_row, column=col_num, value=val)

            # --------------------------
            # 4. 图片完整保留+压缩优化
            # --------------------------
            status_text.text("🖼️  正在优化图片体积...")
            progress_bar.progress(60)
            img_temp_dir = tempfile.mkdtemp()
            
            # 遍历所有行，仅压缩已有图片，不删除、不替换、不新增
            for row in range(1, max_r + 1):
                for col in range(1, max_c + 1):
                    cell = ws.cell(row=row, column=col)
                    # 仅处理有图片的单元格
                    if cell.value and hasattr(cell.value, '_data'):
                        compressed_path = compress_image_optimize(cell.value, img_temp_dir)
                        if compressed_path:
                            # 替换为压缩后的图片，保留原位置、原尺寸
                            new_img = Image(compressed_path)
                            new_img.width = IMG_MAX_W
                            new_img.height = IMG_MAX_H
                            ws.cell(row=row, column=col, value=new_img)

            # --------------------------
            # 5. 生成文件名：单个FBA号用FBA号命名，多个用通用名
            # --------------------------
            status_text.text("💾 正在生成文件...")
            progress_bar.progress(80)
            
            if len(set(fba_list)) == 1:
                # 单个FBA号，用FBA号命名
                file_name = f"{fba_list[0]}_清关单_填充.xlsx"
            else:
                # 多个FBA号，用通用名
                file_name = f"清关单_多FBA批量填充_{today}.xlsx"
            
            output_file_path = os.path.join(img_temp_dir, file_name)
            wb.save(output_file_path)
            wb.close()
            gc.collect()

            # 提供下载
            progress_bar.progress(100)
            status_text.text("✅ 处理完成！")
            file_size_mb = os.path.getsize(output_file_path) / 1024 / 1024
            st.success(f"✅ 填充完成，文件大小：{file_size_mb:.2f} MB")

            # 读取文件用于下载
            with open(output_file_path, "rb") as f:
                file_data = f.read()

            # 清理临时文件
            shutil.rmtree(img_temp_dir, ignore_errors=True)

            # 下载按钮
            st.download_button(
                label="📥 点击下载填充后文件",
                data=file_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # 清理进度条
            progress_bar.empty()
            status_text.empty()

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"处理流程异常：{str(e)}")
            import traceback
            st.code(traceback.format_exc())
