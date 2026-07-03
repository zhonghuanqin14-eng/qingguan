from flask import Flask, request, jsonify, send_file
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
import os, zipfile, tempfile

app = Flask(__name__)

# ========== 固定配置项 ==========
# 固定模板文件名（请确保此文件和app.py放在同一文件夹）
TEMPLATE_FILE_NAME = "AL0-SBU6B5D6EZU6S.xlsx"
# 账号固定配置
ACCOUNT_INFO = {
    "39": {
        "shipper_name": "Shenzhen Longyuan Junjie Technology Co., Ltd.",
        "shipper_addr": "Room 502, No. 5, Ruiyuan Second Lane, Nanlian Community, Longgang Subdistrict, Longgang District, Shenzhen,Guangdong,China",
        "contact": "ZHOUJUNJIE",
        "phone": "+8613427679670"
    },
    "79": {
        "shipper_name": "Mingtongsheng (Shenzhen) E-commerce Co., Ltd.",
        "shipper_addr": "3A Zijing Pavilion, Building 4, Baizhu Garden, 249 Zhuguang Road, Longlian Community, Tao Yuan Street, Nanshan District, Shenzhen,Guangdong,China",
        "contact": "LIMIN",
        "phone": "+8613902478270"
    }
}

# ========== 模板单元格坐标（已根据你的截图精准定位） ==========
CELL_MAP = {
    "fba_no": "J8",               # FBA编号
    # 顶部发货人
    "ship_name": "B7",
    "ship_addr": "B8",
    "ship_contact": "B9",
    "ship_tel": "B10",
    # 顶部进口商（和发货人一致）
    "imp_name": "E7",
    "imp_addr": "E8",
    "imp_contact": "E9",
    "imp_tel": "E10",
    # 底部制造商（已根据你的截图精准定位C38/C39）
    "manu_name": "C38",
    "manu_addr": "C39",
    "data_start_row": 22,         # 明细起始行
    "total_row": 36               # 合计汇总行
}

# 临时文件目录
TMP_DIR = tempfile.gettempdir()
# 模板文件完整路径（和app.py同目录）
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_FILE_NAME)

@app.route("/api/generate", methods=["POST"])
def generate_file():
    try:
        # 校验固定模板文件是否存在
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({"code": 1, "msg": f"固定模板文件不存在！请确保{TEMPLATE_FILE_NAME}和app.py放在同一文件夹"})

        # 接收上传文件与账号
        data_file = request.files["dataFile"]
        acc_id = request.form.get("account")
        acc_data = ACCOUNT_INFO[acc_id]

        # 1. 读取数据源Excel，按FBA分组
        df = pd.read_excel(data_file)
        fba_groups = df.groupby("FBA编号")
        fba_list = list(fba_groups.groups.keys())
        if len(fba_list) == 0:
            return jsonify({"code": 1, "msg": "数据源未读取到FBA编号列！"})

        # 2. 加载固定模板
        tpl_wb = load_workbook(TEMPLATE_PATH)
        tpl_ws = tpl_wb.active

        # 3. 批量生成每个FBA单文件
        out_files = []
        for fba_id, group_df in fba_groups:
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "清关单"

            # ========== 修复1：先复制所有单元格样式（含合并单元格），解决边框丢失问题 ==========
            # 先复制所有单元格的内容和样式，MergedCell只复制样式不赋值
            for row in tpl_ws.iter_rows():
                for cell in row:
                    # 目标单元格
                    new_cell = new_ws.cell(row=cell.row, column=cell.column)
                    # 复制所有样式（边框、字体、填充、对齐等）
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    # 仅非合并单元格赋值，合并单元格跳过赋值
                    if not isinstance(cell, MergedCell):
                        new_cell.value = cell.value

            # 再复制合并单元格范围（确保合并区域正确）
            for mr in tpl_ws.merged_cells.ranges:
                new_ws.merge_cells(str(mr))

            # 复制列宽和行高
            for col_letter in tpl_ws.column_dimensions:
                new_ws.column_dimensions[col_letter].width = tpl_ws.column_dimensions[col_letter].width
            for row_num in tpl_ws.row_dimensions:
                new_ws.row_dimensions[row_num].height = tpl_ws.row_dimensions[row_num].height

            # ========== 修复2：强制写入账号信息，确保制造商信息不丢失 ==========
            # 发货人信息（直接写入固定值）
            new_ws[CELL_MAP["ship_name"]] = acc_data["shipper_name"]
            new_ws[CELL_MAP["ship_addr"]] = acc_data["shipper_addr"]
            new_ws[CELL_MAP["ship_contact"]] = f"Contact:{acc_data['contact']}"
            new_ws[CELL_MAP["ship_tel"]] = f"Phone:{acc_data['phone']}"
            
            # 进口商信息（和发货人完全一致）
            new_ws[CELL_MAP["imp_name"]] = acc_data["shipper_name"]
            new_ws[CELL_MAP["imp_addr"]] = acc_data["shipper_addr"]
            new_ws[CELL_MAP["imp_contact"]] = f"Contact:{acc_data['contact']}"
            new_ws[CELL_MAP["imp_tel"]] = f"Phone:{acc_data['phone']}"
            
            # 制造商信息（强制写入，精准定位C38/C39，解决丢失问题）
            new_ws[CELL_MAP["manu_name"]] = acc_data["shipper_name"]
            new_ws[CELL_MAP["manu_addr"]] = acc_data["shipper_addr"]

            # 填写当前FBA编号
            new_ws[CELL_MAP["fba_no"]] = fba_id

            # 清空模板原有示例明细（仅清空22行开始的明细区域，不影响顶部和底部的固定信息）
            start_r = CELL_MAP["data_start_row"]
            for r in range(start_r, start_r + 100):
                for c in range(2, 17):
                    cell = new_ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = None

            # 写入明细，原产国固定CN
            data_rows = group_df.values.tolist()
            for idx, row_data in enumerate(data_rows):
                curr_r = start_r + idx
                new_ws.cell(row=curr_r, column=2, value=row_data[0])    # 零件号
                new_ws.cell(row=curr_r, column=3, value=row_data[1])    # 品名
                new_ws.cell(row=curr_r, column=4, value=row_data[2])    # 材质
                new_ws.cell(row=curr_r, column=5, value=row_data[3])    # HTS编码
                new_ws.cell(row=curr_r, column=8, value="CN")           # 原产国固定CN
                new_ws.cell(row=curr_r, column=9, value=row_data[7])    # 数量
                new_ws.cell(row=curr_r, column=10, value=row_data[8])   # 单价
                new_ws.cell(row=curr_r, column=11, value=f"=J{curr_r}*I{curr_r}") # 总价公式
                new_ws.cell(row=curr_r, column=13, value=row_data[11])  # 箱数
                new_ws.cell(row=curr_r, column=14, value=row_data[12])  # 毛重
                new_ws.cell(row=curr_r, column=15, value=row_data[13])  # 净重
                new_ws.cell(row=curr_r, column=16, value=row_data[14])  # CBM

                # 复制模板行样式（确保明细行样式和模板一致）
                src_r = CELL_MAP["data_start_row"]
                for c in range(2, 17):
                    src_cell = tpl_ws.cell(row=src_r, column=c)
                    dst_cell = new_ws.cell(row=curr_r, column=c)
                    if not isinstance(src_cell, MergedCell) and not isinstance(dst_cell, MergedCell):
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.alignment = copy(src_cell.alignment)

            # 更新底部合计公式
            data_end_r = start_r + len(data_rows) - 1
            total_r = CELL_MAP["total_row"]
            new_ws.cell(row=total_r, column=11, value=f"=SUM(K{start_r}:K{data_end_r})")
            new_ws.cell(row=total_r, column=13, value=f"=SUM(M{start_r}:M{data_end_r})")
            new_ws.cell(row=total_r, column=14, value=f"=SUM(N{start_r}:N{data_end_r})")
            new_ws.cell(row=total_r, column=15, value=f"=SUM(O{start_r}:O{data_end_r})")
            new_ws.cell(row=total_r, column=16, value=f"=SUM(P{start_r}:P{data_end_r})")

            # 保存单文件
            file_path = os.path.join(TMP_DIR, f"{fba_id}.xlsx")
            new_wb.save(file_path)
            out_files.append(file_path)

        # 打包所有文件为ZIP
        zip_path = os.path.join(TMP_DIR, "FBA清关单批量文件.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for f in out_files:
                zf.write(f, os.path.basename(f))

        return jsonify({
            "code": 0,
            "msg": "生成成功",
            "fileCount": len(out_files),
            "downloadUrl": "/download?zip=" + os.path.basename(zip_path)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"code": 1, "msg": f"处理异常：{str(e)}"})

@app.route("/download")
def download_zip():
    zip_name = request.args.get("zip")
    zip_full = os.path.join(TMP_DIR, zip_name)
    return send_file(zip_full, as_attachment=True)

# 前端页面路由
@app.route("/")
def index_page():
    return send_file("index.html")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)